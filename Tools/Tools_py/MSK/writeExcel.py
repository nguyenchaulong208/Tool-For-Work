import pandas as pd
from openpyxl import load_workbook

def save_data_done_to_excel(data_done, excel_path):
    df = pd.DataFrame(data_done)

    # Load file Excel đang xử lý
    wb = load_workbook(excel_path)

    # Chọn đúng sheet DATA_DONE
    if "DATA_DONE" not in wb.sheetnames:
        raise Exception("Không tìm thấy sheet DATA_DONE trong file Excel!")

    ws = wb["DATA_DONE"]

    # Xóa dữ liệu cũ (giữ lại header)
    ws.delete_rows(2, ws.max_row)

    # Ghi dữ liệu mới vào sheet DATA_DONE
    for row in df.itertuples(index=False):
        ws.append(list(row))

    # Lưu lại file Excel gốc
    wb.save(excel_path)

    print(f"Đã ghi dữ liệu vào sheet DATA_DONE trong file: {excel_path}")
