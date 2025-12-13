import pandas as pd
from io import BytesIO
import streamlit as st
from openpyxl import load_workbook
from form_utils import copy_row_style, write_row_values

def preview_form_sheet(form_file, sheet_name=None):
    """Chuyển sheet form thành DataFrame để hiển thị preview có checkbox."""
    wb = load_workbook(form_file, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    data = list(ws.values)
    df_preview = pd.DataFrame(data)
    # thêm cột checkbox để chọn dòng
    df_preview["Chọn dòng"] = False
    return df_preview

def save_with_form_dynamic_by_index(
    merged,
    form_file,
    output_name,
    sheet_name,
    header_rows,
    body_rows,
    footer_rows,
    body_start_col=1
):
    wb = load_workbook(form_file)
    ws = wb[sheet_name]

    merged = merged.fillna("")
    body_data = merged.values.tolist()
    rows_needed = len(body_data)
    rows_available = len(body_rows)

    # Nếu thiếu dòng → chèn trước dòng đầu footer
    if rows_needed > rows_available:
        insert_at = min(footer_rows) if footer_rows else ws.max_row
        ws.insert_rows(insert_at, amount=rows_needed - rows_available)
        style_model_row = body_rows[0] if body_rows else (header_rows[-1] if header_rows else 1)
        max_col = body_start_col + len(merged.columns) - 1
        for r in range(insert_at, insert_at + (rows_needed - rows_available)):
            copy_row_style(ws, style_model_row, r, max_col)
        body_rows += list(range(insert_at, insert_at + (rows_needed - rows_available)))

    # Nếu thừa dòng → xoá từ cuối body
    elif rows_needed < rows_available:
        rows_to_delete = rows_available - rows_needed
        delete_at = body_rows[-rows_to_delete]
        ws.delete_rows(delete_at, amount=rows_to_delete)
        body_rows = body_rows[:-rows_to_delete]

    # Ghi dữ liệu vào vùng body
    for i, row_values in enumerate(body_data):
        write_row_values(ws, body_rows[i], body_start_col, row_values)

    wb.save(output_name)
    st.success(f"✅ Đã tạo file mới dựa trên form: {output_name}")

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    st.download_button(
        label="📥 Tải file kết quả (giữ header/footer form)",
        data=buffer.getvalue(),
        file_name=output_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )