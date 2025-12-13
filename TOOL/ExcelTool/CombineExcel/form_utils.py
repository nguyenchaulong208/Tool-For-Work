from openpyxl.cell.cell import MergedCell

def copy_row_style(ws, src_row_idx, tgt_row_idx, max_col):
    """Copy style từ một dòng mẫu sang dòng đích."""
    for col in range(1, max_col + 1):
        src_cell = ws.cell(row=src_row_idx, column=col)
        tgt_cell = ws.cell(row=tgt_row_idx, column=col)
        tgt_cell.number_format = src_cell.number_format
        tgt_cell.font = src_cell.font
        tgt_cell.fill = src_cell.fill
        tgt_cell.border = src_cell.border
        tgt_cell.alignment = src_cell.alignment

def write_row_values(ws, row_idx, start_col, values):
    """Ghi giá trị từng ô, xử lý MergedCell bằng cách ghi vào ô gốc của vùng merge."""
    for offset, value in enumerate(values):
        c_idx = start_col + offset
        cell = ws.cell(row=row_idx, column=c_idx)
        if isinstance(cell, MergedCell):
            for merged_range in ws.merged_cells.ranges:
                if (row_idx, c_idx) in merged_range.cells:
                    top_left = merged_range.min_row, merged_range.min_col
                    ws.cell(row=top_left[0], column=top_left[1], value=value)
                    break
        else:
            cell.value = value